from __future__ import annotations

import json
import logging
import shutil
import sys
import zipfile
from pathlib import Path
from typing import Any

import click

from trader_concept import (
    ensure_output_dirs,
    process_traders,
    save_currency_settings,
    save_general_settings,
)
from trader_concept import (
    main as main_gsheets,
)


def _load_config_toml() -> dict[str, Any]:
    import tomllib

    try:
        with open("config.toml", "rb") as f:
            return dict(tomllib.load(f))
    except (FileNotFoundError, tomllib.TOMLDecodeError):
        return {}

logger = logging.getLogger(__name__)

SHEET_NAMES = [
    "Базовые настройки",
    "Настройки состояния",
    "Лицензии",
    "Торговцы",
    "Одежда торговцев",
    "Категории",
    "Товары",
    "Цены",
    "Валюта",
    "Наличка",
]

SHEET_TO_MODEL = {
    "Базовые настройки": "general_settings",
    "Настройки состояния": "accepted_states",
    "Лицензии": "licenses",
    "Торговцы": "traders_raw",
    "Одежда торговцев": "traders_loadouts",
    "Категории": "categories_template",
    "Товары": "products_all",
    "Цены": "prices_raw",
    "Валюта": "currency_types_raw",
    "Наличка": "currencies_raw",
}


# ── Local Input ────────────────────────────────────────────────────────────────


def load_csv_dir(input_dir: str) -> dict[str, list[dict[str, Any]]]:
    """Read all CSV files from a directory, return dict of sheet_name → records."""
    import csv

    result: dict[str, list[dict[str, Any]]] = {}
    input_path = Path(input_dir)

    if not input_path.is_dir():
        raise click.ClickException(f"Папка не найдена: {input_dir}")

    for sheet_name in SHEET_NAMES:
        csv_file = input_path / f"{sheet_name}.csv"
        if not csv_file.exists():
            logger.warning("Файл не найден, пропускаем: %s", csv_file)
            continue
        with open(csv_file, encoding="utf-8-sig") as f:
            reader = csv.DictReader(f)
            rows: list[dict[str, Any]] = []
            for row in reader:
                cleaned = {k.strip(): v.strip() for k, v in row.items() if k}
                rows.append(cleaned)
            result[sheet_name] = rows
            logger.info("Загружено %d записей из %s", len(rows), csv_file.name)

    missing = [n for n in SHEET_NAMES if n not in result]
    if missing:
        logger.warning("Отсутствуют листы: %s", ", ".join(missing))

    return result


def load_excel(filepath: str) -> dict[str, list[dict[str, Any]]]:
    """Read an Excel (.xlsx) file, return dict of sheet_name → records."""
    try:
        import openpyxl
    except ImportError:
        raise click.ClickException(
            "Для работы с Excel нужен пакет openpyxl. Установите: pip install openpyxl"
        )

    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    result: dict[str, list[dict[str, Any]]] = {}

    for sheet_name in SHEET_NAMES:
        if sheet_name not in wb.sheetnames:
            logger.warning("Лист не найден, пропускаем: %s", sheet_name)
            continue
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows or not rows[0]:
            continue
        headers = [str(h).strip() if h is not None else "" for h in rows[0]]
        data_rows: list[dict[str, Any]] = []
        for row in rows[1:]:
            record = {}
            for i, cell in enumerate(row):
                if i < len(headers) and headers[i]:
                    val = cell if cell is not None else ""
                    if isinstance(val, str):
                        val = val.strip()
                    record[headers[i]] = val
            if record:
                data_rows.append(record)
        result[sheet_name] = data_rows
        logger.info("Загружено %d записей из листа %s", len(data_rows), sheet_name)

    wb.close()
    return result


def load_json_dir(input_dir: str) -> dict[str, list[dict[str, Any]]]:
    """Read all JSON files from a directory, return dict of sheet_name → records."""
    result: dict[str, list[dict[str, Any]]] = {}
    input_path = Path(input_dir)

    if not input_path.is_dir():
        raise click.ClickException(f"Папка не найдена: {input_dir}")

    for sheet_name in SHEET_NAMES:
        json_file = input_path / f"{sheet_name}.json"
        if not json_file.exists():
            logger.warning("Файл не найден, пропускаем: %s", json_file)
            continue
        with open(json_file, encoding="utf-8") as f:
            data = json.load(f)
            if isinstance(data, list):
                result[sheet_name] = data
                logger.info("Загружено %d записей из %s", len(data), json_file.name)
            else:
                logger.warning(
                    "Ожидался список записей в %s, получен %s",
                    json_file.name,
                    type(data).__name__,
                )

    missing = [n for n in SHEET_NAMES if n not in result]
    if missing:
        logger.warning("Отсутствуют листы: %s", ", ".join(missing))

    return result


def load_local_data(
    sheet_data: dict[str, list[dict[str, Any]]],
) -> dict[str, Any]:
    """Build the same data dict as load_data() from local files instead of gspread."""
    from models import input as input_models

    general_settings = input_models.GeneralSettings.from_raw(
        sheet_data.get("Базовые настройки", [])
    )
    accepted_states = input_models.AcceptedStates.from_raw(
        sheet_data.get("Настройки состояния", [])
    )

    raw_licenses = sheet_data.get("Лицензии", [])
    licenses = [input_models.License.from_raw(item) for item in raw_licenses]

    raw_traders = sheet_data.get("Торговцы", [])
    traders_raw = [input_models.Trader.from_raw(item) for item in raw_traders]
    trader_names = [t.givenName for t in traders_raw]

    raw_loadouts = sheet_data.get("Одежда торговцев", [])
    traders_loadouts = [input_models.Loadout.from_raw(item) for item in raw_loadouts]

    raw_categories = sheet_data.get("Категории", [])
    categories_template = [input_models.Category.from_raw(item) for item in raw_categories]

    raw_products = sheet_data.get("Товары", [])
    products_all = [input_models.Product.from_raw(item, trader_names) for item in raw_products]

    raw_currency_types = sheet_data.get("Валюта", [])
    currency_types_raw = [
        input_models.CurrencyType.from_raw(item)
        for item in raw_currency_types
    ]

    raw_currencies = sheet_data.get("Наличка", [])
    currencies_raw = [input_models.Currency.from_raw(item) for item in raw_currencies]

    loadouts_by_trader: dict[int, list[input_models.Loadout]] = {}
    for lo in traders_loadouts:
        loadouts_by_trader.setdefault(lo.id, []).append(lo)

    raw_prices = sheet_data.get("Цены", [])
    prices_map: dict[str, input_models.TraderPrices] = {}
    for p in raw_prices:
        price_obj = input_models.TraderPrices.from_raw(p, trader_names)
        prices_map[price_obj.className] = price_obj

    return {
        "general_settings": general_settings,
        "accepted_states": accepted_states,
        "licenses": licenses,
        "traders_raw": traders_raw,
        "traders_loadouts": traders_loadouts,
        "loadouts_by_trader": loadouts_by_trader,
        "categories_template": categories_template,
        "products_all": products_all,
        "currency_types_raw": currency_types_raw,
        "currencies_raw": currencies_raw,
        "prices_map": prices_map,
    }


# ── ZIP ────────────────────────────────────────────────────────────────────────


def pack_output(output_dir: str, output_zip: str) -> str:
    """Pack the generated files into a ZIP archive."""
    output_path = Path(output_dir)
    if not output_path.is_dir():
        raise click.ClickException(f"Папка не найдена: {output_dir}")

    zip_path = Path(output_zip)
    zip_path.parent.mkdir(parents=True, exist_ok=True)

    with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:
        for file_path in output_path.rglob("*"):
            if file_path.is_file():
                arcname = str(file_path.relative_to(output_path.parent))
                zf.write(file_path, arcname)

    logger.info("Архив создан: %s", zip_path)
    return str(zip_path)


# ── CLI ────────────────────────────────────────────────────────────────────────


@click.group()
def cli() -> None:
    """Генератор конфигов TraderX для DayZ."""
    logging.basicConfig(
        level=logging.INFO,
        format="%(levelname)s: %(message)s",
        stream=sys.stderr,
    )


@cli.command()
@click.option("--sheet-id", "-s", help="ID Google Sheets таблицы")
@click.option("--credentials", "-c", default="credentials.json", help="Путь к credentials.json")
@click.option("--output", "-o", default="output/profiles/TraderX", help="Папка для результатов")
@click.option("--zip", "output_zip", help="Упаковать результат в ZIP (путь к файлу)")
@click.option("--cluster/--no-cluster", default=False, help="Кластеризация торговцев по ассортименту (ML)")
def google(sheet_id: str | None, credentials: str, output: str, output_zip: str | None, cluster: bool) -> None:
    """Сгенерировать конфиги из Google Sheets."""
    if not sheet_id:
        import tomllib
        try:
            with open("config.toml", "rb") as f:
                cfg = tomllib.load(f)
            sheet_id = cfg.get("google", {}).get("spreadsheet_id")
            credentials = cfg.get("google", {}).get("credentials_file", credentials)
            output = cfg.get("output_dir", output)
        except (FileNotFoundError, tomllib.TOMLDecodeError):
            pass

    if not sheet_id:
        raise click.ClickException(
            "Укажите ID таблицы через --sheet-id или в config.toml"
        )

    data = main_gsheets(
        credentials_file=credentials,
        spreadsheet_id=sheet_id,
        output_dir=output,
    )

    if cluster and data:
        _run_clustering(data=data)

    if output_zip:
        pack_output(output, output_zip)


@cli.command()
@click.argument("input", type=click.Path(exists=True))
@click.option("--format", "-f", "input_format",
              type=click.Choice(["auto", "csv", "json", "xlsx"]), default="auto")
@click.option("--output", "-o", default="output/profiles/TraderX", help="Папка для результатов")
@click.option("--zip", "output_zip", help="Упаковать результат в ZIP (путь к файлу)")
@click.option("--cluster/--no-cluster", default=False, help="Кластеризация торговцев по ассортименту (ML)")
def local(input: str, input_format: str, output: str, output_zip: str | None, cluster: bool) -> None:
    """Сгенерировать конфиги из локальных файлов (CSV/JSON/Excel)."""
    input_path = Path(input)

    if input_format == "auto":
        if input_path.is_dir():
            if list(input_path.glob("*.csv")):
                input_format = "csv"
            elif list(input_path.glob("*.json")):
                input_format = "json"
            else:
                raise click.ClickException(
                    "В папке не найдены CSV или JSON файлы. "
                    "Имена файлов должны соответствовать названиям листов."
                )
        elif input_path.suffix.lower() == ".xlsx":
            input_format = "xlsx"
        else:
            raise click.ClickException(
                "Формат не определён. Используйте --format csv/json/xlsx"
            )

    logger.info("Загрузка данных из %s (формат: %s)", input, input_format)

    if input_format == "csv":
        sheet_data = load_csv_dir(input)
    elif input_format == "json":
        sheet_data = load_json_dir(input)
    elif input_format == "xlsx":
        sheet_data = load_excel(input)
    else:
        raise click.ClickException(f"Неподдерживаемый формат: {input_format}")

    if not sheet_data:
        raise click.ClickException("Не удалось загрузить ни одного листа")

    data = load_local_data(sheet_data)
    ensure_output_dirs(output)

    if cluster:
        _run_clustering(data=data)

    output_traders = process_traders(
        traders_raw=data["traders_raw"],
        products_all=data["products_all"],
        prices_map=data["prices_map"],
        categories_template=data["categories_template"],
        loadouts_by_trader=data["loadouts_by_trader"],
        output_dir=output,
    )

    save_general_settings(
        general_settings=data["general_settings"],
        accepted_states=data["accepted_states"],
        licenses=data["licenses"],
        output_traders=output_traders,
        output_dir=output,
    )

    save_currency_settings(
        currency_types_raw=data["currency_types_raw"],
        currencies_raw=data["currencies_raw"],
        output_dir=output,
    )

    logger.info("Готово! Все файлы созданы в папке %s", output)

    if output_zip:
        pack_output(output, output_zip)


@cli.command()
@click.argument("source_dir", type=click.Path(exists=True, file_okay=False))
@click.argument("output_zip", type=click.Path())
def pack(source_dir: str, output_zip: str) -> None:
    """Упаковать готовые файлы в ZIP архив."""
    pack_output(source_dir, output_zip)


@cli.command()
def init_config() -> None:
    """Создать config.toml с настройками по умолчанию."""
    src = Path(__file__).parent / "config.toml"
    dst = Path("config.toml")
    if dst.exists():
        click.echo("config.toml уже существует")
        return
    if src.exists():
        shutil.copy2(src, dst)
        click.echo("Создан config.toml")
    else:
        click.echo("Не найден шаблон config.toml", err=True)


@cli.command()
@click.option("--port", "-p", default=8080, show_default=True, help="Порт веб-сервера")
@click.option("--host", default="0.0.0.0", show_default=True, help="Хост веб-сервера")
def web(port: int, host: str) -> None:
    """Запустить веб-интерфейс генератора."""
    from web.app import run_web
    run_web(host=host, port=port)


@cli.command()
@click.option(
    "--output-dir",
    default="output/profiles/TraderX",
    help="Корневая папка сгенерированных конфигов",
    show_default=True,
)
@click.option(
    "--local-cache",
    help="Локальная папка со скачанными Stock JSON (если уже есть, без FTP)",
)
@click.option(
    "--ftp-host",
    help="FTP сервер хостинга (user@host или просто host)",
)
@click.option("--ftp-password", default="", help="FTP пароль (если не указан — запросит)")
@click.option(
    "--ftp-remote-path",
    default="/Server/profiles/TraderX/TraderXDatabase/Stock",
    help="Путь к Stock на FTP сервере",
    show_default=True,
)
@click.option(
    "--ftp-tls/--no-ftp-tls",
    default=False,
    help="Использовать FTPS (FTP over TLS)",
)
@click.option(
    "--threshold", default=2.0, show_default=True, help="Z-score порог для выбросов"
)
def stock_report(
    output_dir: str,
    local_cache: str | None,
    ftp_host: str | None,
    ftp_password: str,
    ftp_remote_path: str,
    ftp_tls: bool,
    threshold: float,
) -> None:
    """Анализ состояния складов: спрос, мёртвый сток, аномалии.

    Примеры:

      python -m cli stock-report --ftp-host user@185.207.214.77

      python -m cli stock-report --local-cache /home/u/stock_backup

      python -m cli stock-report  (сгенерированные файлы, все на 100%)
    """
    try:
        from ml_stock_analytics import (
            analyze_stock,
            fetch_stock_via_ftp,
            load_product_info,
            load_stock_files,
            load_traders,
            print_stock_report,
        )
    except ImportError:
        logger.exception("Не удалось загрузить ml_stock_analytics")
        sys.exit(1)

    ftp_user: str = "anonymous"

    if ftp_host:
        if "@" in ftp_host:
            ftp_user, ftp_host = ftp_host.rsplit("@", 1)
        if not ftp_password:
            ftp_password = click.prompt(
                f"FTP пароль для {ftp_user}@{ftp_host}",
                hide_input=True,
            )
        logger.info("Загрузка стока с FTP %s ...", ftp_host)
        local_cache = fetch_stock_via_ftp(
            host=ftp_host,
            user=ftp_user,
            password=ftp_password,
            remote_path=ftp_remote_path,
            use_tls=ftp_tls,
        )

    if local_cache is None:
        local_cache = str(Path(output_dir) / "TraderXDatabase" / "Stock")

    products_dir = str(Path(output_dir) / "TraderXConfig" / "Products")
    general_settings = str(Path(output_dir) / "TraderXConfig" / "TraderXGeneralSettings.json")

    stock_data = load_stock_files(local_cache)
    if not stock_data:
        logger.error("Не найдены stock файлы в %s", local_cache)
        sys.exit(1)

    product_info = load_product_info(products_dir)
    if not product_info:
        logger.error("Не найдены product файлы в %s", products_dir)
        sys.exit(1)

    traders = load_traders(general_settings)

    logger.info(
        "Загружено stock: %d, products: %d, traders: %d",
        len(stock_data), len(product_info), len(traders),
    )

    result = analyze_stock(stock_data, product_info, traders, outlier_threshold=threshold)
    print_stock_report(result)


@cli.command()
@click.option(
    "--output-dir",
    default="output/profiles/TraderX",
    help="Корневая папка сгенерированных конфигов",
    show_default=True,
)
@click.option(
    "--loot-path",
    default="output/profiles/lootzones.json",
    help="Путь к lootzones.json",
    show_default=True,
)
@click.option(
    "--balance/--no-balance",
    default=False,
    help="Авто-балансировка цен на основе тира редкости + stockSettings",
)
@click.option(
    "--apply/--dry-run",
    default=False,
    help="--apply = перезаписать файлы, --dry-run = только показать изменения",
)
@click.option(
    "--target-sellprice",
    default="50,200,800,3000,10000,30000",
    help="Целевые sellPrice для Tier0-5 через запятую",
    show_default=True,
)
@click.option(
    "--stock-demand-map",
    default=None,
    help="Кастомные коэф. спроса для stockSettings: ss=factor,ss=factor (напр. 0=0.3,98=1.5)",
    show_default=True,
)
@click.option(
    "--to-sheet/--no-sheet",
    default=False,
    help="Записать отчёты в Google Sheets (читает credentials из config.toml)",
)
def price_report(
    output_dir: str,
    loot_path: str,
    balance: bool,
    apply: bool,
    target_sellprice: str,
    stock_demand_map: str | None,
    to_sheet: bool = False,
) -> None:
    """Анализ экономики: цены лута, редкость, дисбаланс.

    Анализирует lootzones.json + Product файлы, выводит:
      - Tier-распределение sellPrice (что игрок получает за продажу лута)
      - Аномально дорогие/дешёвые предметы в каждом тире
      - Обзор рынка снаряжения (stockSettings=0)
      - Коэффициенты «сколько лута продать, чтобы купить геар»

    С --balance анализирует + предлагает авто-балансировку цен.
    С --balance --apply перезаписывает цены в Product файлах.
    """
    try:
        from ml_price_analysis import (
            analyze_economy,
            auto_balance_prices,
            print_price_report,
        )
    except ImportError:
        logger.exception("Не удалось загрузить ml_price_analysis")
        sys.exit(1)

    products_dir = str(Path(output_dir) / "TraderXConfig" / "Products")

    parsed_target: dict[int, int] = {}
    parts = target_sellprice.split(",")
    for i, val in enumerate(parts[:6]):
        try:
            parsed_target[i] = int(val.strip())
        except ValueError:
            logger.warning("Неверное значение target_sellprice[%d]: %s", i, val)

    parsed_demand_map: dict[int, float] | None = None
    if stock_demand_map:
        parsed_demand_map = {}
        for pair in stock_demand_map.split(","):
            if "=" not in pair:
                continue
            try:
                k, v = pair.split("=", 1)
                parsed_demand_map[int(k.strip())] = float(v.strip())
            except (ValueError, TypeError):
                logger.warning("Неверное значение stock-demand-map: %s", pair)

    logger.info("Анализ экономики: loot=%s products=%s", loot_path, products_dir)
    result = analyze_economy(loot_path=loot_path, products_dir=products_dir)
    print_price_report(result)

    balance_stats = None
    if balance:
        balance_stats = auto_balance_prices(
            products_dir=products_dir,
            loot_path=loot_path,
            dry_run=not apply,
            target_sellprice=parsed_target or None,
            stock_demand_map=parsed_demand_map,
        )

    if to_sheet:
        cfg = _load_config_toml()
        google_cfg = cfg.get("google", {})
        credentials_file: str = google_cfg.get("credentials_file", "credentials.json")
        spreadsheet_id: str = google_cfg.get("spreadsheet_id", "")

        if not spreadsheet_id:
            logger.error("spreadsheet_id не найден в config.toml [google]")
            sys.exit(1)

        try:
            from ml_sheet_reports import write_all_economy_reports
            from trader_concept import create_gspread_client, open_sheet

            client = create_gspread_client(credentials_file)
            sheet = open_sheet(client, spreadsheet_id)
            write_all_economy_reports(sheet, result, balance_stats)
        except Exception:
            logger.exception("Не удалось записать отчёты в Google Sheets")
            sys.exit(1)


@cli.command("attachments-report")
@click.option(
    "--source",
    type=click.Choice(["google", "local"]),
    default="google",
    show_default=True,
    help="Источник данных: google (Sheets) или local (файлы)",
)
@click.option(
    "--data-dir",
    default=None,
    type=click.Path(exists=True),
    help="Путь к папке/файлу с CSV/JSON/Excel (при --source local)",
)
@click.option(
    "--sheet-id",
    "-s",
    default=None,
    help="ID Google Sheets таблицы",
)
@click.option(
    "--credentials",
    "-c",
    default="credentials.json",
    help="Путь к credentials.json",
)
def attachments_report(
    source: str,
    data_dir: str | None,
    sheet_id: str | None,
    credentials: str,
) -> None:
    """Анализ отсутствующих обвесов для оружия.

    Сценарий 1: обвес определён у оружия, но не продаётся ни одним трейдером.
    Сценарий 2: трейдер продаёт оружие, но не продаёт обвес для него.

    Примеры:

      trader attachments-report --source google

      trader attachments-report --source local ./data
    """
    try:
        from ml_attachment_analytics import (
            analyze_missing_attachments,
            print_attachment_report,
        )
    except ImportError:
        logger.exception("Не удалось загрузить ml_attachment_analytics")
        sys.exit(1)

    products_all: list = []
    traders_raw: list = []

    if source == "google":
        cfg = _load_config_toml()
        google_cfg = cfg.get("google", {})
        credentials = credentials or google_cfg.get("credentials_file", "credentials.json")
        sheet_id = sheet_id or google_cfg.get("spreadsheet_id", "")

        if not sheet_id:
            raise click.ClickException(
                "Укажите ID таблицы через --sheet-id или в config.toml"
            )

        try:
            from trader_concept import create_gspread_client, load_data, open_sheet

            client = create_gspread_client(credentials)
            sheet = open_sheet(client, sheet_id)
            trader_names: list[str] = []
            data = load_data(sheet, trader_names)
            products_all = data["products_all"]
            traders_raw = data["traders_raw"]
        except Exception:
            logger.exception("Не удалось загрузить данные из Google Sheets")
            sys.exit(1)
    else:
        if not data_dir:
            raise click.ClickException("Укажите --data-dir для local-источника")

        input_path = Path(data_dir)
        if input_path.is_dir():
            if list(input_path.glob("*.csv")):
                sheet_data = load_csv_dir(data_dir)
            elif list(input_path.glob("*.json")):
                sheet_data = load_json_dir(data_dir)
            else:
                raise click.ClickException(
                    "В папке не найдены CSV или JSON файлы"
                )
        elif input_path.suffix.lower() == ".xlsx":
            sheet_data = load_excel(data_dir)
        else:
            raise click.ClickException(f"Неподдерживаемый формат: {input_path.suffix}")

        data = load_local_data(sheet_data)
        products_all = data["products_all"]
        traders_raw = data["traders_raw"]

    logger.info(
        "Анализ обвесов: товаров=%d, трейдеров=%d",
        len(products_all), len(traders_raw),
    )

    result = analyze_missing_attachments(products_all, traders_raw)
    print_attachment_report(result)


@cli.command()
@click.option(
    "--output-dir",
    default="output/profiles/TraderX",
    help="Корневая папка сгенерированных конфигов",
    show_default=True,
)
@click.option(
    "--min-profit",
    default=100,
    help="Минимальная прибыль в рублях для обнаружения",
    show_default=True,
)
@click.option(
    "--min-profit-pct",
    default=1.0,
    help="Минимальная маржа в процентах",
    show_default=True,
)
@click.option(
    "--repair-cost-pct",
    default=0.3,
    help="Оценка стоимости ремонта в долях от buyPrice (0.3 = 30%%)",
    show_default=True,
)
@click.option(
    "--attachment/--no-attachment",
    default=True,
    help="Анализировать арбитраж обвеса",
)
@click.option(
    "--repair/--no-repair",
    default=True,
    help="Анализировать арбитраж ремонта",
)
@click.option(
    "--multi-hop/--no-multi-hop",
    default=False,
    help="Поиск цепочек из 3+ торговцев (DFS)",
)
@click.option(
    "--max-chain-length",
    default=5,
    help="Максимальная длина цепочки multi-hop",
    show_default=True,
)
@click.option(
    "--to-sheet/--no-sheet",
    default=False,
    help="Записать отчёты в Google Sheets",
)
def arbitrage(
    output_dir: str,
    min_profit: int,
    min_profit_pct: float,
    repair_cost_pct: float,
    attachment: bool,
    repair: bool,
    multi_hop: bool,
    max_chain_length: int,
    to_sheet: bool = False,
) -> None:
    """Поиск экономических дыр и арбитража между торговцами.

    Анализирует цены между торговцами на одинаковые предметы,
    обвес (дизассемблирование/сборка), ремонт изношенных предметов.

    Примеры:

      trader arbitrage

      trader arbitrage --multi-hop --min-profit 1000

      trader arbitrage --no-attachment --no-repair --to-sheet
    """
    try:
        from ml_arbitrage_detector import (
            detect_arbitrages,
            print_arbitrage_report,
        )
    except ImportError:
        logger.exception("Не удалось загрузить ml_arbitrage_detector")
        sys.exit(1)

    products_dir = str(Path(output_dir) / "TraderXConfig" / "Products")
    settings_path = str(Path(output_dir) / "TraderXConfig" / "TraderXGeneralSettings.json")

    logger.info("Поиск арбитража: products=%s settings=%s", products_dir, settings_path)
    result = detect_arbitrages(
        products_dir=products_dir,
        general_settings_path=settings_path,
        repair_cost_pct=repair_cost_pct,
        min_profit_abs=min_profit,
        min_profit_pct=min_profit_pct,
        enable_cross_trader=True,
        enable_attachment=attachment,
        enable_repair=repair,
        enable_multi_hop=multi_hop,
        max_chain_length=max_chain_length,
    )
    print_arbitrage_report(result)

    if to_sheet:
        cfg = _load_config_toml()
        google_cfg = cfg.get("google", {})
        credentials_file: str = google_cfg.get("credentials_file", "credentials.json")
        spreadsheet_id: str = google_cfg.get("spreadsheet_id", "")

        if not spreadsheet_id:
            logger.error("spreadsheet_id не найден в config.toml [google]")
            sys.exit(1)

        try:
            from ml_sheet_reports import write_arbitrage_report
            from trader_concept import create_gspread_client, open_sheet

            client = create_gspread_client(credentials_file)
            sheet = open_sheet(client, spreadsheet_id)
            write_arbitrage_report(sheet, result)
        except Exception:
            logger.exception("Не удалось записать отчёт в Google Sheets")
            sys.exit(1)

def _run_clustering(
    data: dict[str, Any] | None = None,
    sheet_id: str | None = None,
    credentials_file: str = "",
) -> None:
    """Run trader clustering analysis."""
    try:
        from ml_trader_clustering import cluster_traders, print_cluster_report
    except ImportError:
        logger.error(
            "Для кластеризации требуется scikit-learn. Установите: pip install '.[ml]'"
        )
        return

    cluster_names: dict[str, str] = {}
    cfg = _load_config_toml()
    names_cfg = cfg.get("ml", {}).get("names", {})
    if isinstance(names_cfg, dict):
        cluster_names = {str(k): str(v) for k, v in names_cfg.items()}

    if data is None:
        from trader_concept import create_gspread_client, load_data, open_sheet

        try:
            client = create_gspread_client(credentials_file)
            sheet = open_sheet(client, sheet_id or "")
            trader_names: list[str] = []
            data = load_data(sheet, trader_names)
        except Exception:
            logger.exception("Не удалось загрузить данные из Google Sheets для кластеризации")
            return

    traders_raw = data.get("traders_raw", [])
    products_all = data.get("products_all", [])
    categories_template = data.get("categories_template", [])

    if not traders_raw:
        logger.warning("Нет данных о торговцах для кластеризации")
        return

    result = cluster_traders(
        traders_raw, products_all, categories_template, cluster_names=cluster_names or None,
    )
    print_cluster_report(result)


if __name__ == "__main__":
    cli()
